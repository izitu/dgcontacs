import openpyxl

def tel_mody(tel):
    tel = tel.replace(' ','')
    tel = tel.replace('-', '')
    tel = tel.replace('(', '')
    tel = tel.replace(')', '')
    if len(tel) == 12 : tel = '+7' + tel[2:]
    if len(tel) == 11: tel = '+7' + tel[1:]
    if len(tel) > 12: tel = ' --- ' + tel + ' --- '
    return tel


# читаем excel-файл
wb = openpyxl.load_workbook('30 набор контакты.xlsx')


# получаем активный лист
sheet = wb.active
my_file = open("import-rodit2.csv", "w", encoding='utf-8')

str_csv = 'Name,' +\
          'Given Name,' \
          'Family Name,' \
          'Group Membership,' \
          'E-mail 1 - Type, E-mail 1 - Value,' \
          'Phone 1 - Value,' \
          'Organization 1 - Name,' \
          'Organization 1 - Title,' \
          'Custom Field 1 - Type,' \
          'Custom Field 1 - Value,' \
          'Custom Field 2 - Type,' \
          'Custom Field 2 - Value,' \
          'Custom Field 3 - Type,' \
          'Custom Field 3 - Value,' \
          'Custom Field 4 - Type,' \
          'Custom Field 4 - Value,'\
          'Custom Field 5 - Type,' \
          'Custom Field 5 - Value,' \
          'Custom Field 6 - Type,' \
          'Custom Field 6 - Value,' \
          'Custom Field 7 - Type,' \
          'Custom Field 7 - Value \n'
my_file.write(str_csv)

rows = sheet.max_row
cols = sheet.max_column
for i in range(3, rows + 1):
    fio = str(sheet.cell(row=i, column=8).value).split()
    tel = str(sheet.cell(row=i, column=9).value)
    mail = str(sheet.cell(row=i, column=10).value)
    rodfio1 = str(sheet.cell(row=i, column=1).value)
    rodtel1 = str(sheet.cell(row=i, column=2).value)
    mail1 = str(sheet.cell(row=i, column=3).value)
    rodfio2 = str(sheet.cell(row=i, column=5).value)
    rodtel2 = str(sheet.cell(row=i, column=6).value)
    mail2 = str(sheet.cell(row=i, column=7).value)
    metro = str(sheet.cell(row=i, column=11).value)


    print(fio)
    if fio[0] == 'None':
        continue
    if len(mail) == 4  :
        mail = ','
    else:
        mail = '*,' + mail

    all_str = fio[0] + ' ' + fio[1] + ',' + \
              fio[0] + ',' + \
              fio[1] + ',' + \
              'ДГ ::: * myContacts,' + \
              mail + ',' + \
              tel_mody(tel) + ',' + \
              'ДГ' +  ',' + \
              'Родитель' + ',' + \
              'Ребенок' + ',' + rodfio1 + ',' + \
              'Телефон Ребенка' + ',' + tel_mody(rodtel1) + ',' + \
              'Email Ребенка' + ',' + mail1 + ',' + \
              'Родитель 1' + ',' + rodfio2 + ',' + \
              'Телефон Р1' + ',' + tel_mody(rodtel2) + ',' + \
              'Email Р1' + ',' + mail2 + ',' + \
              'Метро' + ',' + metro + '\n'
    all_str = all_str.replace('None', '')
    print(all_str)
    my_file.write(all_str)

my_file.close()
