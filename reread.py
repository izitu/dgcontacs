import openpyxl

# читаем excel-файл
wb = openpyxl.load_workbook('30 набор контакты.xlsx')

# печатаем список листов

# получаем активный лист
sheet = wb.active

# печатаем значение ячейки A1
print(sheet['A5'].value)
# печатаем значение ячейки B1
print(sheet['B5'].value)
print(sheet['C5'].value)

rows = sheet.max_row
cols = sheet.max_column

for i in range(1, rows + 1):
    string = ''
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        if (cell.value)==None:
            strr = ''
        else:
            strr = str(cell.value)
        string = string + strr + ','
    print(string)